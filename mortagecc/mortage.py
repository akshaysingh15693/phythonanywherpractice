# -*- coding: utf-8 -*-
"""
Created on Tue May 19 22:32:34 2020

@author: Spartan
"""
import openpyxl
import pandas as pd
import numpy  as np
from datetime import datetime

path = 'C:/Users/Spartan/Desktop/'
attachment=path+'dataset.xlsx'
excel_path=str(attachment)
wb=openpyxl.load_workbook(excel_path)
sheet=wb['input_pool']
df=pd.read_excel(excel_path)

class Mortage():
    
    def __init__(self,path):
        print(path)
        self.cc()

        
    def cc(self):
        global i
        i=2  
        wals=0.4
        lastr=len(df)
        lastra=lastr+1
        ffb=0.12
        cp=1
        while(i<=lastra):
#            ab=sheet.cell(row=i,column=3).value
            ota=self.factorc()
            psa=self.py_sk_tm()
            ga=self.gca()
            shrta=self.shotpa()
            ffi=min(1,(ffb*ota),(ffb*psa),(ffb*ga),(ffb*shrta))
            print('ffi is',ffi)
            sumab=df['adjusted.bal'].sum()
            print('sumab is',sumab)
            sumffi=sumab*ffi
            waff=min(1,((sumffi/sumab)*shrta))
            print('waff is',waff)
            cc=(waff*wals)*100
            print('cc is',cc)
            sheet.cell(row=i, column=16).value = cc
            i=i+1
        wb.save(excel_path)
        wb.close()
             
         

    def factorc(self): 
#        print(path)
#        path = 'C:/Users/Spartan/Desktop/'
#        attachment=path+'dataset.xlsx'
#        excel_path=str(attachment)
#        wb=openpyxl.load_workbook(excel_path)
#        sheet=wb['input_pool']
        ot=sheet.cell(row=i,column=11).value
        print(ot)
#        ota=s_o_a(a)
#        print(ota)
        if(ot=='sh'):
            ota=1.2
        elif(ot=='btl'):
            ota=1.3
        else:
             ota=1
        print('ota is',ota)
        return ota
        
    def py_sk_tm(self):
#        attachment=path+'dataset.xlsx'
#        excel_path=str(attachment)
        isd=sheet.cell(row=i,column=9).value
#        isd2 = datetime.strptime(isd, '%m-%b-%y')
        isd1=isd.date()
#        iisd=int(isd1)
        con='1970-01-01'
        asofdt='2020-03-31'
        con2=datetime.strptime(con,'%Y-%m-%d')
        con1=con2.date()
        print('con1 is',con1)
        asofdt2=datetime.strptime(asofdt,'%Y-%m-%y')
        asofdt1=asofdt2.date()
#        iasofdt=int(asofdt1)
        print('asofdt1 is ',asofdt1)
        print('isd1 is ',isd1)
        frm= asofdt1-isd1
        frm1=frm.days
        frm2=((frm1/365)*12)
        print('frm is ',frm2)
        if((isd1 != con1 ) and (frm2>6)):
            pyskt="pastgr6m"
        elif((isd1 != con1 ) and (frm2<=6)):
            pyskt="le6m"
        else:
            pyskt="missing"
        ip=sheet.cell(row=i,column=8).value
        if(ip=="discount" and pyskt=="le6m"):
            psa=1.5
        else:
            psa=1
        print('psa is',psa)
        return psa
    
    def gca(self):
        gk= df.groupby('region',as_index=False).sum()
        reg=sheet.cell(row=i,column=12).value
        rab=df['adjusted.bal'].sum()
        abc=gk[gk.region==reg]
        ab=abc['adjusted.bal']
        iab=int(ab)
        geoc=(iab/rab)*100
        if(geoc>20):
            ga=1.5
        else:
            ga=1
        print('ga is ',ga)
        return ga
    
    def shotpa(self):
        lc=len(df)
        nl=np.log(lc)
        lgc=16.0839/nl
        if(lc<250):
            shrta=lgc
        elif(lc==1):
            shrta=40
        elif(lc>250):
            shrta=1
        print('shrta is',shrta)
        return shrta
        
        
        
        
app=Mortage('C:/Users/Spartan/Desktop/')
#app.factorc('C:/Users/Spartan/Desktop/')
        

        